"""
ACQUISITION EDGE - Deal Proposal Generator
Generates investor-ready financial proposals for business acquisitions

Usage:
    python deal_proposal_generator.py deal_config.json

Configuration JSON structure in comments at end of file
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DealProposalGenerator:
    def __init__(self, config):
        self.config = config
        self.wb = Workbook()
        
        # Brand colors - Signal Gold on Obsidian
        self.colors = {
            'header_bg': "1A1A1A",
            'header_text': "C9A84C",
            'subheader_bg': "2D2D2D",
            'success_green': "00B050",
            'input_blue': "0000FF"
        }
        
        # Standard styles
        self.header_fill = PatternFill(start_color=self.colors['header_bg'], 
                                       end_color=self.colors['header_bg'], fill_type="solid")
        self.header_font = Font(color=self.colors['header_text'], bold=True, size=14, name="Arial")
        self.subheader_fill = PatternFill(start_color=self.colors['subheader_bg'], 
                                          end_color=self.colors['subheader_bg'], fill_type="solid")
        self.subheader_font = Font(color="FFFFFF", bold=True, size=11, name="Arial")
        self.label_font = Font(bold=True, size=10, name="Arial")
        self.value_font = Font(size=10, name="Arial")
        self.blue_input_font = Font(color=self.colors['input_blue'], size=10, name="Arial")

    def create_summary_sheet(self):
        ws = self.wb.active
        ws.title = "Deal Summary"
        
        deal = self.config['deal_overview']
        
        # Title section
        ws['A1'] = "ACQUISITION EDGE"
        ws['A1'].font = Font(color=self.colors['header_text'], bold=True, size=18, name="Arial")
        ws['A2'] = "Deal Structure Proposal"
        ws['A2'].font = Font(size=12, name="Arial")
        ws['A3'] = deal['company_name']
        ws['A3'].font = Font(bold=True, size=14, name="Arial")
        ws['A4'] = f"Generated: {deal.get('date', 'N/A')}"
        ws['A4'].font = Font(size=9, italic=True, name="Arial")
        
        # Deal Overview
        ws['A6'] = "DEAL OVERVIEW"
        ws['A6'].fill = self.header_fill
        ws['A6'].font = self.header_font
        ws.merge_cells('A6:D6')
        
        ws['A7'] = "Purchase Price"
        ws['B7'] = deal['purchase_price']
        ws['B7'].number_format = '$#,##0'
        ws['B7'].font = self.blue_input_font
        
        ws['C7'] = "Asset Value (FF&E + Inventory)"
        ws['D7'] = deal['asset_value']
        ws['D7'].number_format = '$#,##0'
        ws['D7'].font = self.blue_input_font
        
        ws['A8'] = "Closing Costs / Working Capital"
        ws['B8'] = deal['closing_costs']
        ws['B8'].number_format = '$#,##0'
        ws['B8'].font = self.blue_input_font
        
        ws['C8'] = "Asset Arbitrage Discount"
        ws['D8'] = '=(D7-B7)/D7'
        ws['D8'].number_format = '0.0%'
        
        ws['A9'] = "Total Capital Required"
        ws['B9'] = '=B7+B8'
        ws['B9'].number_format = '$#,##0'
        ws['B9'].font = Font(bold=True, size=10, name="Arial")
        
        if deal.get('real_estate_option'):
            ws['C9'] = "Real Estate Option Strike"
            ws['D9'] = deal['real_estate_option']
            ws['D9'].number_format = '$#,##0'
            ws['D9'].font = self.blue_input_font
        
        # Capital Stack
        ws['A11'] = "CAPITAL STACK"
        ws['A11'].fill = self.header_fill
        ws['A11'].font = self.header_font
        ws.merge_cells('A11:D11')
        
        headers = ['Source', 'Amount', 'Terms', 'Security']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=12, column=idx)
            cell.value = header
            cell.fill = self.subheader_fill
            cell.font = self.subheader_font
        
        row = 13
        total_formula_parts = []
        for item in self.config['capital_stack']:
            ws.cell(row=row, column=1).value = item['source']
            ws.cell(row=row, column=2).value = item['amount']
            ws.cell(row=row, column=2).number_format = '$#,##0'
            ws.cell(row=row, column=2).font = self.blue_input_font
            ws.cell(row=row, column=3).value = item['terms']
            ws.cell(row=row, column=4).value = item['security']
            total_formula_parts.append(f'B{row}')
            row += 1
        
        ws.cell(row=row, column=1).value = "TOTAL RAISE"
        ws.cell(row=row, column=2).value = f'=SUM({",".join(total_formula_parts)})'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True, size=11, name="Arial")
        
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 24

    def create_cashflow_sheet(self):
        ws = self.wb.create_sheet("Cash Flow Analysis")
        
        ops = self.config['operations']
        
        ws['A1'] = "CASH FLOW ANALYSIS"
        ws['A1'].fill = self.header_fill
        ws['A1'].font = self.header_font
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "OPERATING ASSUMPTIONS"
        ws['A3'].fill = self.subheader_fill
        ws['A3'].font = self.subheader_font
        ws.merge_cells('A3:B3')
        
        ws['A4'] = "Current SDE"
        ws['B4'] = ops['current_sde']
        ws['B4'].number_format = '$#,##0'
        ws['B4'].font = self.blue_input_font
        
        ws['A5'] = ops.get('manager_label', 'Operations Manager Salary')
        ws['B5'] = ops.get('manager_salary', 0)
        ws['B5'].number_format = '$#,##0'
        ws['B5'].font = self.blue_input_font
        
        ws['A6'] = "Adjusted EBITDA"
        ws['B6'] = '=B4-B5'
        ws['B6'].number_format = '$#,##0'
        ws['B6'].font = Font(bold=True, size=10, name="Arial")
        
        ws['A8'] = "ANNUAL DEBT SERVICE"
        ws['A8'].fill = self.subheader_fill
        ws['A8'].font = self.subheader_font
        ws.merge_cells('A8:B8')
        
        row = 9
        debt_cells = []
        for debt in self.config['debt_service']:
            ws.cell(row=row, column=1).value = debt['label']
            ws.cell(row=row, column=2).value = debt['annual_amount']
            ws.cell(row=row, column=2).number_format = '$#,##0'
            ws.cell(row=row, column=2).font = self.blue_input_font
            debt_cells.append(f'B{row}')
            row += 1
        
        ws.cell(row=row, column=1).value = "Total Annual Obligations"
        ws.cell(row=row, column=2).value = f'=SUM({",".join(debt_cells)})'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        total_obligations_row = row
        row += 2
        
        ws.cell(row=row, column=1).value = "NET DISTRIBUTABLE CASH FLOW"
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        ws.cell(row=row, column=1).value = "Annual EBITDA"
        ws.cell(row=row, column=2).value = '=B6'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1
        
        ws.cell(row=row, column=1).value = "Less: Total Obligations"
        ws.cell(row=row, column=2).value = f'=-B{total_obligations_row}'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1
        
        ws.cell(row=row, column=1).value = "Net Distributable Cash Flow"
        ws.cell(row=row, column=2).value = f'=B{row-2}+B{row-1}'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True, size=11, name="Arial")
        net_cf_row = row
        row += 2
        
        dist = self.config.get('distributions', {'gp_pct': 0.40, 'lp_pct': 0.60})
        
        ws.cell(row=row, column=1).value = "DISTRIBUTION SPLIT"
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        ws.cell(row=row, column=1).value = f"GP Share ({dist['gp_pct']*100:.0f}%)"
        ws.cell(row=row, column=2).value = f'=B{net_cf_row}*{dist["gp_pct"]}'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        gp_annual_row = row
        row += 1
        
        ws.cell(row=row, column=1).value = f"LP Share ({dist['lp_pct']*100:.0f}%)"
        ws.cell(row=row, column=2).value = f'=B{net_cf_row}*{dist["lp_pct"]}'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        lp_annual_row = row
        row += 2
        
        ws.cell(row=row, column=1).value = "MONTHLY CASH FLOW"
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        ws.cell(row=row, column=1).value = "GP Monthly Income"
        ws.cell(row=row, column=2).value = f'=B{gp_annual_row}/12'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        row += 1
        
        ws.cell(row=row, column=1).value = "LP Monthly Distributions"
        ws.cell(row=row, column=2).value = f'=B{lp_annual_row}/12'
        ws.cell(row=row, column=2).number_format = '$#,##0'
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 16

    def create_exit_sheet(self):
        if 'exit_strategy' not in self.config:
            return
        
        ws = self.wb.create_sheet("Exit Waterfall")
        exit_data = self.config['exit_strategy']
        
        ws['A1'] = f"EXIT WATERFALL ANALYSIS - YEAR {exit_data.get('exit_year', 7)}"
        ws['A1'].fill = self.header_fill
        ws['A1'].font = self.header_font
        ws.merge_cells('A1:C1')
        
        ws['A3'] = "EXIT ASSUMPTIONS"
        ws['A3'].fill = self.subheader_fill
        ws['A3'].font = self.subheader_font
        ws.merge_cells('A3:C3')
        
        row = 4
        ws.cell(row=row, column=1).value = "Business Valuation"
        ws.cell(row=row, column=2).value = "Multiple"
        ws.cell(row=row, column=3).value = "Value"
        row += 1
        
        ws.cell(row=row, column=1).value = "Projected EBITDA"
        ws.cell(row=row, column=2).value = exit_data['business_ebitda']
        ws.cell(row=row, column=2).number_format = '$#,##0'
        ws.cell(row=row, column=2).font = self.blue_input_font
        ebitda_row = row
        row += 1
        
        ws.cell(row=row, column=1).value = "EBITDA Multiple"
        ws.cell(row=row, column=2).value = exit_data['ebitda_multiple']
        ws.cell(row=row, column=2).number_format = '0.0x'
        ws.cell(row=row, column=2).font = self.blue_input_font
        multiple_row = row
        row += 1
        
        ws.cell(row=row, column=1).value = "Business Value"
        ws.cell(row=row, column=3).value = f'=B{ebitda_row}*B{multiple_row}'
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=10, name="Arial")
        biz_value_row = row
        row += 2
        
        if exit_data.get('real_estate_value'):
            ws.cell(row=row, column=1).value = "Real Estate Valuation"
            row += 1
            
            ws.cell(row=row, column=1).value = "Current Value"
            ws.cell(row=row, column=2).value = exit_data['real_estate_current']
            ws.cell(row=row, column=2).number_format = '$#,##0'
            ws.cell(row=row, column=2).font = self.blue_input_font
            re_current_row = row
            row += 1
            
            ws.cell(row=row, column=1).value = "Annual Appreciation"
            ws.cell(row=row, column=2).value = exit_data.get('re_appreciation', 0.04)
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=2).font = self.blue_input_font
            appreciation_row = row
            row += 1
            
            ws.cell(row=row, column=1).value = "Years"
            ws.cell(row=row, column=2).value = exit_data.get('exit_year', 7)
            ws.cell(row=row, column=2).font = self.blue_input_font
            years_row = row
            row += 1
            
            ws.cell(row=row, column=1).value = "Projected RE Value"
            ws.cell(row=row, column=3).value = f'=B{re_current_row}*(1+B{appreciation_row})^B{years_row}'
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True, size=10, name="Arial")
            re_value_row = row
            row += 2
            
            ws.cell(row=row, column=1).value = "TOTAL EXIT VALUE"
            ws.cell(row=row, column=3).value = f'=C{biz_value_row}+C{re_value_row}'
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True, size=12, color=self.colors['header_text'], name="Arial")
            total_exit_row = row
        else:
            ws.cell(row=row, column=1).value = "TOTAL EXIT VALUE"
            ws.cell(row=row, column=3).value = f'=C{biz_value_row}'
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True, size=12, color=self.colors['header_text'], name="Arial")
            total_exit_row = row
        
        row += 2
        ws.cell(row=row, column=1).value = "WATERFALL DISTRIBUTION"
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws.cell(row=row, column=1).value = "Priority"
        ws.cell(row=row, column=2).value = "Payee"
        ws.cell(row=row, column=3).value = "Amount"
        for col in [1, 2, 3]:
            ws.cell(row=row, column=col).fill = self.subheader_fill
            ws.cell(row=row, column=col).font = self.subheader_font
        row += 1
        
        priority = 1
        waterfall_deductions = []
        for item in exit_data.get('waterfall', []):
            ws.cell(row=row, column=1).value = str(priority)
            ws.cell(row=row, column=2).value = item['payee']
            ws.cell(row=row, column=3).value = item['amount']
            ws.cell(row=row, column=3).number_format = '$#,##0'
            if priority == 1:
                ws.cell(row=row, column=3).font = self.blue_input_font
            waterfall_deductions.append(f'C{row}')
            priority += 1
            row += 1
        
        ws.cell(row=row, column=1).value = str(priority)
        ws.cell(row=row, column=2).value = "Remaining Proceeds"
        ws.cell(row=row, column=3).value = f'=C{total_exit_row}-{"-".join(waterfall_deductions)}'
        ws.cell(row=row, column=3).number_format = '$#,##0'
        remaining_row = row
        row += 2
        
        split = exit_data.get('final_split', {'gp_pct': 0.70, 'lp_pct': 0.30})
        
        ws.cell(row=row, column=1).value = "SPLIT OF REMAINING PROCEEDS"
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws.cell(row=row, column=1).value = f"GP ({split['gp_pct']*100:.0f}%)"
        ws.cell(row=row, column=3).value = f'=C{remaining_row}*{split["gp_pct"]}'
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True, size=11, color=self.colors['success_green'], name="Arial")
        row += 1
        
        ws.cell(row=row, column=1).value = f"LP ({split['lp_pct']*100:.0f}%)"
        ws.cell(row=row, column=3).value = f'=C{remaining_row}*{split["lp_pct"]}'
        ws.cell(row=row, column=3).number_format = '$#,##0'
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18

    def generate(self, output_filename):
        self.create_summary_sheet()
        self.create_cashflow_sheet()
        self.create_exit_sheet()
        self.wb.save(output_filename)
        return output_filename

def main():
    if len(sys.argv) < 2:
        print("Usage: python deal_proposal_generator.py deal_config.json")
        sys.exit(1)
    
    with open(sys.argv[1], 'r') as f:
        config = json.load(f)
    
    generator = DealProposalGenerator(config)
    output_file = config.get('output_filename', 'Deal_Proposal.xlsx')
    
    result = generator.generate(output_file)
    print(f"Generated: {result}")

"""
CONFIGURATION JSON TEMPLATE:

{
  "deal_overview": {
    "company_name": "Example Business LLC",
    "date": "4/16/2026",
    "purchase_price": 300000,
    "asset_value": 433000,
    "closing_costs": 25000,
    "real_estate_option": 2375000
  },
  "capital_stack": [
    {
      "source": "Earnest Money (Gator)",
      "amount": 1000,
      "terms": "$2,000 flat return at close",
      "security": "Unsecured"
    },
    {
      "source": "Acquisition Fee (GP)",
      "amount": 6000,
      "terms": "2% fee paid at closing",
      "security": "N/A"
    },
    {
      "source": "Private Money Lender (Tier 1)",
      "amount": 150000,
      "terms": "12% IO, 24-month balloon",
      "security": "1st lien on FF&E ($206k)"
    },
    {
      "source": "Equity Partners (Tier 2)",
      "amount": 175000,
      "terms": "10% pref + 60% equity",
      "security": "60% LLC ownership"
    }
  ],
  "operations": {
    "current_sde": 195664,
    "manager_salary": 70000,
    "manager_label": "Operations Manager Salary"
  },
  "debt_service": [
    {
      "label": "PML Debt Service (12% on $150k)",
      "annual_amount": 18000
    },
    {
      "label": "Investor Preferred Return (10% on $175k)",
      "annual_amount": 17500
    }
  ],
  "distributions": {
    "gp_pct": 0.40,
    "lp_pct": 0.60
  },
  "exit_strategy": {
    "exit_year": 7,
    "business_ebitda": 200000,
    "ebitda_multiple": 4.0,
    "real_estate_current": 2375000,
    "re_appreciation": 0.04,
    "real_estate_value": true,
    "waterfall": [
      {
        "payee": "SBA Senior Debt Payoff",
        "amount": 2100000
      },
      {
        "payee": "Return Equity Partner Capital",
        "amount": 175000
      }
    ],
    "final_split": {
      "gp_pct": 0.70,
      "lp_pct": 0.30
    }
  },
  "output_filename": "Deal_Proposal.xlsx"
}
"""

if __name__ == "__main__":
    main()
